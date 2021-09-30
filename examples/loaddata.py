#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl


def load_xlsx(fl_input, sheet = "Sheet1", splitchain = True):
    # Load the spreadsheet...
    bk = openpyxl.load_workbook(fl_input, data_only = True)
    st = bk[sheet]

    # Fetch entries from the 2nd row...
    entries = []
    for i in st.iter_rows(min_row = 2):
        # Unpack all values
        val = [ j.value for j in i ]

        # Stop iteration if row starts with None
        if val[0] is None: break

        # Unpack parameters
        select, pdb, chain, species = val[:4]

        # Skip unselected...
        if select == "no": 
            print(f"{pdb}-{chain}-{species} is not selected.")
            continue

        # Go through each chain and record...
        if splitchain:
            for eachchain in chain.split(): 
                # Replace the chain column with a single chain identifier...
                val[2] = eachchain
                entries.append(val.copy())
        else:
            entries.append(val.copy())

    return entries




def load_gpcrdb_xlsx(fl_input, sheet = "Sheet1", splitchain = True):
    # Load the spreadsheet...
    bk = openpyxl.load_workbook(fl_input, data_only = True)
    st = bk[sheet]

    # Fetch entries from the 2nd row...
    entries = []
    for i in st.iter_rows(min_row = 2):
        # Unpack all values
        val = [ j.value for j in i ]

        # Stop iteration if row starts with None
        if val[0] is None: break

        # Unpack parameters
        select = val[0]
        pdb    = val[7]
        chain  = val[10]

        # Skip unselected...
        if select == "no": 
            print(f"{pdb}-{chain} is not selected.")
            continue

        # Go through each chain and record...
        if splitchain:
            for eachchain in chain.split(): 
                # Replace the chain column with a single chain identifier...
                val[10] = eachchain
                entries.append(val.copy())
        else:
            entries.append(val.copy())

    return entries




def load_md_xlsx(fl_input, sheet = "Sheet1"):
    # Load the spreadsheet...
    bk = openpyxl.load_workbook(fl_input, data_only = True)
    st = bk[sheet]

    # Fetch entries from the 2nd row...
    entries = []
    for i in st.iter_rows(min_row = 2):
        # Unpack all values
        val = [ j.value for j in i ]

        # Stop iteration if row starts with None
        if val[0] is None: break

        # Unpack parameters
        select, pdb, chain, frame_b, frame_e, frame_s = val[:6]

        # Skip unselected...
        if select == "no": 
            print(f"{pdb}-{chain} is not selected.")
            continue


        entries.append(val.copy())

    return entries




def load_uniprot_species(fl_species):
    spec_dict= {}
    fl_species = "organism.dat"
    with open(fl_species,'r') as fh:
        for line in fh.readlines():
            if line.startswith('#'): continue

            line = line.lower()
            offset = 2
            code = line[ offset + 0  : offset +  5 ]
            ncs  = line[ offset + 17 : ].rstrip()
            k, v = ncs[0], ncs[2:]

            # If there is a new species...
            if len(code.strip()): 
                spec_dict[code] = { k : v }
                code_prev = code
            else: spec_dict[code_prev][k] = v


    return spec_dict




def name_to_code(spec_dict, name = "C"):
    name_to_code_dict = {}
    for k, v in spec_dict.items():
        if not name in v: continue

        k_new = v[name]
        v_new = k

        if not k_new in name_to_code_dict: name_to_code_dict[k_new] = v_new
        else: name_to_code_dict[k_new] += f" | {v_new}"

    return name_to_code_dict
